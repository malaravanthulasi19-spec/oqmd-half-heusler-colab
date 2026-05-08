from pathlib import Path
import pandas as pd

from .strategic_classifier import (
    KNOWN_REPORTED_MATERIALS,
    classify_literature_status,
    load_prior_material_evidence,
)
try:
    from openpyxl.styles import Alignment, Font, PatternFill
except ModuleNotFoundError:
    Alignment = Font = PatternFill = None


DECISION_PRIORITY = [
    "RECOMMENDED_NOVEL_CANDIDATE",
    "STRONG_NOVEL_CANDIDATE_MANUAL_VERIFY",
    "NOVEL_BUT_MODERATE_STABILITY",
    "EXPENSIVE_RARE_BACKUP",
    "NOVEL_BUT_LOW_STABILITY",
    "PRIOR_RUN_CONFLICT_MANUAL_REVIEW",
    "AMBIGUOUS_PRIOR_WORK_REVIEW",
    "REPORTED_DFT_DEFER",
    "DEEP_PRIOR_STUDY_DEFER",
    "TOXICITY_REVIEW_DEFER",
    "RADIOACTIVE_DEFER",
    "HIGHLY_IMPRACTICAL_DEFER",
    "INCOMPLETE_SEARCH_RETRY",
    "LOWER_PRIORITY_REVIEW",
]
FINAL_DECISION_SIMPLE_PRIORITY = ["BEST_NOVEL_CANDIDATE","GOOD_NOVEL_CANDIDATE","MANUAL_REVIEW_REQUIRED","BACKUP_ONLY","LOW_STABILITY_REJECT","TOXICITY_REJECT","REPORTED_REJECT"]

def _to_float(value, default=0.0):
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def make_paper_link(best_url, best_doi) -> str:
    if pd.notna(best_url) and str(best_url).strip():
        return str(best_url).strip()
    if pd.notna(best_doi) and str(best_doi).strip():
        doi = str(best_doi).strip()
        return doi if doi.startswith("http") else f"https://doi.org/{doi}"
    return ""


def _compute_priority_scores(row: dict) -> dict:
    ls = row.get("literature_status", "")
    rel = row.get("evidence_reliability_tier", "")
    unreported = 0
    if ls in {"HIGH_CONFIDENCE_NOT_FOUND", "MEDIUM_CONFIDENCE_NOT_FOUND"}: unreported += 60
    if row.get("novelty_confidence_tier") == "HIGH_CONFIDENCE_UNREPORTED": unreported += 20
    if not bool(row.get("formula_level_evidence_found")): unreported += 10
    if _to_float(row.get("exact_formula_hit_count"), 0) == 0: unreported += 5
    if _to_float(row.get("dft_formula_hit_count"), 0) == 0: unreported += 5
    unreported -= {"REPORTED_DFT":100,"DEEP_PRIOR_STUDY":90,"REPORTED_NON_DFT":80,"AMBIGUOUS_PRIOR_WORK":50,"INCOMPLETE_SEARCH":40}.get(ls,0)
    if bool(row.get("prior_conflict_flag")): unreported -= 70
    if rel == "PRIOR_CONFLICT_REVIEW": unreported -= 60
    unreported = max(0, min(100, unreported))
    st = _to_float(row.get("Stability"), None)
    stability = 30 if st is None else (100 if st<=0.05 else 90 if st<=0.10 else 70 if st<=0.20 else 50 if st<=0.30 else 20 if st<=0.60 else 0)
    bg = _to_float(row.get("Band Gap (eV)"), None)
    band = 30 if bg is None else (100 if 0.5<=bg<=1.5 else 90 if 0.1<=bg<0.5 else 85 if 1.5<bg<=2.0 else 70 if 2.0<bg<=3.0 else 40 if 3.0<bg<=5.0 else 20 if bg==0 else 0)
    practicality = {"PRACTICAL_PRIORITY":100,"EXPENSIVE_RARE_REVIEW":75,"TOXICITY_REVIEW":40,"RADIOACTIVE_REVIEW":20,"HIGHLY_IMPRACTICAL":0}.get(row.get("practicality_tier",""),0)
    penalty=0
    if ls=="REPORTED_DFT": penalty+=100
    if bool(row.get("prior_conflict_flag")): penalty+=95
    if ls=="DEEP_PRIOR_STUDY": penalty+=90
    if _to_float(row.get("keypaper_depth_score"),0)>=80: penalty+=80
    if _to_float(row.get("reported_depth_score"),0)>=75: penalty+=75
    if ls=="REPORTED_NON_DFT": penalty+=60
    if ls=="AMBIGUOUS_PRIOR_WORK": penalty+=50
    if _to_float(row.get("exact_formula_hit_count"),0)>0: penalty+=40
    if _to_float(row.get("dft_formula_hit_count"),0)>0: penalty+=40
    if str(row.get("best_paper_title","")).strip(): penalty+=30
    if str(row.get("best_url","")).strip() or str(row.get("best_doi","")).strip(): penalty+=30
    penalty=max(0,min(100,penalty))
    final=max(0,min(100,0.45*unreported+0.30*stability+0.15*band+0.10*practicality-0.35*penalty))
    return {"unreported_priority_score":unreported,"stability_priority_score":stability,"bandgap_priority_score":band,"practicality_priority_score":practicality,"prior_literature_penalty":penalty,"final_priority_order_score":final,"priority_order_reason":"Unreported > Stability > Band gap > Practicality > Prior literature penalty"}


def assign_material_decision_status(row: dict) -> dict:
    material = str(row.get("Material", "") or "").strip()
    prior = load_prior_material_evidence(None, material)
    merged = {**row, **prior, **classify_literature_status({**row, **prior})}
    if bool(merged.get("prior_conflict_flag")) or material in KNOWN_REPORTED_MATERIALS:
        merged["prior_conflict_flag"] = True
    if merged.get("prior_conflict_flag") and not merged.get("prior_best_doi") and material in KNOWN_REPORTED_MATERIALS:
        merged["prior_best_doi"] = KNOWN_REPORTED_MATERIALS[material].get("doi", "")
        merged["prior_best_url"] = KNOWN_REPORTED_MATERIALS[material].get("url", "")
    if merged.get("prior_conflict_flag") or material in KNOWN_REPORTED_MATERIALS:
        rel = {"evidence_reliability_tier":"PRIOR_CONFLICT_REVIEW","evidence_reliability_reason":"Known/prior reported conflict"}
    elif bool(merged.get("source_error")):
        rel = {"evidence_reliability_tier":"INCOMPLETE_RETRY","evidence_reliability_reason":"Source error"}
    elif bool(merged.get("google_scholar_checked")) and bool(merged.get("openalex_checked")) and bool(merged.get("semantic_scholar_checked")) and (not bool(merged.get("formula_level_evidence_found"))) and _to_float(merged.get("exact_formula_hit_count"),0)==0 and _to_float(merged.get("dft_formula_hit_count"),0)==0:
        rel = {"evidence_reliability_tier":"COMPLETE_HIGH_RELIABILITY","evidence_reliability_reason":"Complete checked sources, no formula-level evidence"}
    elif bool(merged.get("google_scholar_checked")) and bool(merged.get("openalex_checked")) and merged.get("best_evidence_match_type")=="element_system_weak":
        rel = {"evidence_reliability_tier":"COMPLETE_WITH_WEAK_HITS","evidence_reliability_reason":"Only weak element-system evidence"}
    else:
        rel = {"evidence_reliability_tier":"LOW_RELIABILITY","evidence_reliability_reason":"Fallback"}
    merged.update(rel)
    merged.update(_compute_priority_scores(merged))
    ls=merged["literature_status"]; pr=merged.get("practicality_tier",""); st=_to_float(merged.get("Stability"),999); bg=_to_float(merged.get("Band Gap (eV)"),0)
    link = make_paper_link(merged.get("prior_best_url") or merged.get("best_url",""), merged.get("prior_best_doi") or merged.get("best_doi",""))
    if ls=="REPORTED_DFT": d=("REPORTED_DFT_DEFER","Use as reference/comparison only")
    elif ls=="DEEP_PRIOR_STUDY": d=("DEEP_PRIOR_STUDY_DEFER","Do not claim novelty without manual citation review")
    elif bool(merged.get("prior_conflict_flag")) or material in KNOWN_REPORTED_MATERIALS: d=("PRIOR_RUN_CONFLICT_MANUAL_REVIEW","Do not claim novelty until prior evidence is checked")
    elif ls=="AMBIGUOUS_PRIOR_WORK": d=("AMBIGUOUS_PRIOR_WORK_REVIEW","Manual literature check required")
    elif ls=="INCOMPLETE_SEARCH": d=("INCOMPLETE_SEARCH_RETRY","Rerun search before using this material")
    elif pr=="HIGHLY_IMPRACTICAL": d=("HIGHLY_IMPRACTICAL_DEFER","Do not prioritize")
    elif pr=="RADIOACTIVE_REVIEW": d=("RADIOACTIVE_DEFER","Do not prioritize")
    elif pr=="TOXICITY_REVIEW": d=("TOXICITY_REVIEW_DEFER","Avoid for first paper candidate")
    elif ls=="HIGH_CONFIDENCE_NOT_FOUND" and merged["evidence_reliability_tier"] in {"COMPLETE_HIGH_RELIABILITY","COMPLETE_WITH_WEAK_HITS"} and merged["unreported_priority_score"]>=90 and st<=0.10 and 0.1<=bg<=2.0 and pr=="PRACTICAL_PRIORITY" and merged["prior_literature_penalty"]<30 and not bool(merged.get("prior_conflict_flag")): d=("RECOMMENDED_NOVEL_CANDIDATE","Top candidate for manual confirmation and DFT study")
    elif ls in {"HIGH_CONFIDENCE_NOT_FOUND","MEDIUM_CONFIDENCE_NOT_FOUND"} and merged["unreported_priority_score"]>=80 and st<=0.20 and 0.1<=bg<=3.0 and pr in {"PRACTICAL_PRIORITY","EXPENSIVE_RARE_REVIEW"} and merged["prior_literature_penalty"]<50: d=("STRONG_NOVEL_CANDIDATE_MANUAL_VERIFY","Manual literature check recommended")
    elif pr=="EXPENSIVE_RARE_REVIEW": d=("EXPENSIVE_RARE_BACKUP","Keep as backup, not first choice")
    elif ls in {"HIGH_CONFIDENCE_NOT_FOUND","MEDIUM_CONFIDENCE_NOT_FOUND"} and 0.10 < st <=0.30: d=("NOVEL_BUT_MODERATE_STABILITY","Use as backup candidate")
    elif ls in {"HIGH_CONFIDENCE_NOT_FOUND","MEDIUM_CONFIDENCE_NOT_FOUND"} and st>0.30: d=("NOVEL_BUT_LOW_STABILITY","Lower priority unless metastability is acceptable")
    elif bg>5: d=("NOVEL_BUT_APPLICATION_MISMATCH","Not ideal for thermoelectric/spintronic priority")
    else: d=("LOWER_PRIORITY_REVIEW","Keep for later review")
    return {**merged,"material_decision_status":d[0],"material_decision_reason":d[1],"final_recommendation":d[1],"paper_link_if_reported":link}


def _style_sheet(ws):
    if not Font or not Alignment:
        return
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for c in ws[1]:
        c.font = Font(bold=True)
    wrap_cols = {"T", "U", "V", "W", "X"}
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        letter = col[0].column_letter
        ws.column_dimensions[letter].width = min(max_len + 2, 60)
        if letter in wrap_cols:
            for cell in col:
                cell.alignment = Alignment(wrap_text=True, vertical="top")


def dedupe_dataframe_columns(df: pd.DataFrame) -> pd.DataFrame:
    return df.loc[:, ~df.columns.duplicated()].copy()


def export_material_screening_master(rows, hits, coverage, output_dir) -> Path:
    output_dir = Path(output_dir)
    out_path = output_dir / "13_material_screening_master.xlsx"
    base_df = pd.DataFrame(rows).copy()
    decision_records = [assign_material_decision_status(r) for r in base_df.to_dict(orient="records")]
    df = pd.DataFrame(decision_records)
    df = dedupe_dataframe_columns(df)
    df["_decision"] = pd.Categorical(df["material_decision_status"], categories=DECISION_PRIORITY, ordered=True)
    for col in ["unreported_priority_score", "stability_priority_score", "bandgap_priority_score", "practicality_priority_score", "prior_literature_penalty", "final_priority_order_score"]:
        if col not in df.columns:
            df[col] = 0
    final_decision = df.sort_values(["_decision", "unreported_priority_score", "stability_priority_score", "bandgap_priority_score", "practicality_priority_score", "prior_literature_penalty", "final_priority_order_score"], ascending=[True, False, False, False, False, True, False]).drop(columns=["_decision"])
    final_decision = dedupe_dataframe_columns(final_decision)
    final_decision = _build_simple_ranked_list(final_decision)
    final_decision = dedupe_dataframe_columns(final_decision)
    fd_cols = ["Rank","Material","Band Gap (eV)","Stability","Formation Energy / ΔE","Space Group","Prototype","Automated Status","material_decision_status","literature_status","evidence_reliability_tier","unreported_priority_score","stability_priority_score","bandgap_priority_score","practicality_priority_score","prior_literature_penalty","final_priority_order_score","priority_order_reason","prior_conflict_flag","prior_reported_status","prior_best_paper_title","prior_best_doi","prior_best_url","Unreported Confidence Score","novelty_confidence_tier","reported_depth_score","keypaper_depth_score","final_selection_score","final_material_priority_tier","practicality_tier","radioactive_elements","highly_toxic_elements","expensive_rare_elements","best_paper_title","paper_link_if_reported","material_decision_reason","final_recommendation","reviewer_notes"]
    top10_cols = ["Rank","Material","Composition","Band Gap (eV)","Formation Energy / ΔE","Stability","Prototype","Space Group","OQMD Entry ID","Automated Status","Reason","material_decision_status","material_decision_reason","final_recommendation","paper_link_if_reported","Unreported Confidence Score","Reported Evidence Score","novelty_confidence_tier","formula_level_evidence_found","exact_formula_hit_count","dft_formula_hit_count","reported_depth_score","reported_depth_tier","keypaper_depth_score","keypaper_depth_tier","keypaper_context_groups_detected","novelty_score","half_heusler_validity_score","stability_score","practicality_score","application_score","literature_risk_penalty","metadata_quality_score","final_selection_score","final_material_priority_tier","practicality_tier","input_half_heusler_verified","literature_half_heusler_context_found","half_heusler_filter_status","best_paper_title","best_doi","best_url","reviewer_notes"]
    for cols in [fd_cols, top10_cols]:
        for c in cols:
            if c not in final_decision.columns:
                final_decision[c] = ""
    try:
        writer = pd.ExcelWriter(out_path, engine="openpyxl")
    except ModuleNotFoundError:
        out_path.write_bytes(b"")
        return out_path
    with writer as xw:
        simple_cols = ["Final Rank","Material","Unreported Score","Unreported Status","Stability","Stability Grade","Toxicity / Practicality","Band Gap (eV)","Band Gap Grade","Reported Paper Link","Final Decision","Final Reason"]
        final_ranked = dedupe_dataframe_columns(final_decision[simple_cols])
        final_ranked.to_excel(xw, sheet_name="Final_Ranked_List", index=False)
        final_decision = dedupe_dataframe_columns(final_decision)
        final_decision[fd_cols].to_excel(xw, sheet_name="Final_Decision", index=False)
        final_decision[final_decision["material_decision_status"] == "RECOMMENDED_NOVEL_CANDIDATE"][fd_cols].to_excel(xw, "Recommended_Novel", index=False)
        final_decision[final_decision["material_decision_status"].isin(["STRONG_NOVEL_CANDIDATE_MANUAL_VERIFY","PRIOR_RUN_CONFLICT_MANUAL_REVIEW","AMBIGUOUS_PRIOR_WORK_REVIEW"])][fd_cols].to_excel(xw, "Strong_Manual_Verify", index=False)
        final_decision[final_decision["material_decision_status"].isin(["NOVEL_BUT_MODERATE_STABILITY","EXPENSIVE_RARE_BACKUP","NOVEL_BUT_LOW_STABILITY","LOWER_PRIORITY_REVIEW"])][fd_cols].to_excel(xw, "Backup_Candidates", index=False)
        final_decision[final_decision["material_decision_status"].isin(["REPORTED_DFT_DEFER","DEEP_PRIOR_STUDY_DEFER","TOXICITY_REVIEW_DEFER","RADIOACTIVE_DEFER","HIGHLY_IMPRACTICAL_DEFER","INCOMPLETE_SEARCH_RETRY","NOVEL_BUT_APPLICATION_MISMATCH"])][fd_cols].to_excel(xw, "Reported_or_Defer", index=False)
        dedupe_dataframe_columns(final_decision[top10_cols]).to_excel(xw, "All_Top10", index=False)
        (pd.DataFrame(hits) if hits else pd.DataFrame([{"note":"No hit audit rows available for this run."}])).to_excel(xw, "All_Hits_Audit", index=False)
        (pd.DataFrame(coverage) if coverage else pd.DataFrame([{"note":"No search coverage rows available for this run."}])).to_excel(xw, "Search_Coverage", index=False)
        pd.DataFrame([{"note":"Search strategy details are retained in pipeline logs and audit sheets."}]).to_excel(xw, "Search_Strategy", index=False)
        validation_df = _build_validation_checks(final_decision)
        validation_df.to_excel(xw, "Validation_Checks", index=False)
        pd.DataFrame({"Legend":["Use Recommended_Novel first.","Automated Status = literature search status.","literature_status = strategic interpretation of literature evidence.","evidence_reliability_tier = confidence in the search outcome.","material_decision_status = final material-selection decision.","Ranking priority is: 1. unreported novelty, 2. stability, 3. band gap, 4. practicality, 5. prior literature penalty.","Do not use PRIOR_RUN_CONFLICT_MANUAL_REVIEW, REPORTED_DFT_DEFER, or DEEP_PRIOR_STUDY_DEFER as novelty candidates."]}).to_excel(xw, "Legend", index=False)
        wb = xw.book
        fill_map = {"RECOMMENDED_NOVEL_CANDIDATE":"C6EFCE","STRONG_NOVEL_CANDIDATE_MANUAL_VERIFY":"E2F0D9","NOVEL_BUT_MODERATE_STABILITY":"FFF2CC","EXPENSIVE_RARE_BACKUP":"FFF2CC","NOVEL_BUT_LOW_STABILITY":"FFF2CC","AMBIGUOUS_PRIOR_WORK_REVIEW":"FCE4D6"}
        red = "F8CBAD"
        for ws in wb.worksheets:
            _style_sheet(ws)
            if ws.title in {"Final_Ranked_List","Final_Decision","Recommended_Novel","Strong_Manual_Verify","Backup_Candidates","Reported_or_Defer","All_Top10"}:
                headers = [c.value for c in ws[1]]
                key_col = "material_decision_status" if "material_decision_status" in headers else ("Final Decision" if "Final Decision" in headers else None)
                if key_col:
                    ci = headers.index(key_col) + 1
                    for r in range(2, ws.max_row + 1):
                        v = ws.cell(r, ci).value
                        color = fill_map.get(v, red if str(v).endswith("DEFER") or v in {"INCOMPLETE_SEARCH_RETRY"} else None)
                        if key_col == "Final Decision":
                            color = {
                                "BEST_NOVEL_CANDIDATE": "C6EFCE",
                                "GOOD_NOVEL_CANDIDATE": "E2F0D9",
                                "MANUAL_REVIEW_REQUIRED": "FCE4D6",
                                "BACKUP_ONLY": "FFF2CC",
                                "LOW_STABILITY_REJECT": "F8CBAD",
                                "TOXICITY_REJECT": "EA9999",
                                "REPORTED_REJECT": "D9D2E9",
                            }.get(v)
                        if color and PatternFill:
                            ws.cell(r, ci).fill = PatternFill(fill_type="solid", fgColor=color)
    return out_path


def _series(df: pd.DataFrame, col: str, default=""):
    if col not in df.columns:
        return pd.Series([default] * len(df), index=df.index)
    value = df.loc[:, col]
    if isinstance(value, pd.DataFrame):
        value = value.iloc[:, 0]
    return value


def _build_simple_ranked_list(df: pd.DataFrame) -> pd.DataFrame:
    out = dedupe_dataframe_columns(df.copy())
    ls = out.get("literature_status", pd.Series([""] * len(out))).fillna("")
    exact = _series(out, "exact_formula_evidence_count", None)
    if exact is None or exact.isna().all():
        exact = _series(out, "exact_formula_hit_count", 0)
    exact = pd.to_numeric(exact, errors="coerce").fillna(0)
    perm = pd.to_numeric(_series(out, "permutation_formula_evidence_count", 0), errors="coerce").fillna(0)
    deep = _series(out, "deep_dft_property_evidence_count", None)
    if deep is None or deep.isna().all():
        deep = _series(out, "dft_formula_hit_count", 0)
    deep = pd.to_numeric(deep, errors="coerce").fillna(0)
    fam = pd.to_numeric(_series(out, "family_level_evidence_count", 0), errors="coerce").fillna(0)
    conflict = _series(out, "prior_conflict_flag", False).fillna(False).astype(bool)
    known = _series(out, "known_reported_composition_flag", False).fillna(False).astype(bool)
    reported_mask = ls.isin(["REPORTED_DFT", "REPORTED_NON_DFT", "DEEP_PRIOR_STUDY"]) | (exact > 0) | (perm > 0) | (deep > 0)
    out["Unreported Score"] = 80
    out.loc[reported_mask, "Unreported Score"] = 0
    out.loc[conflict | known, "Unreported Score"] = 20
    out.loc[(fam > 0) & ~reported_mask & ~(conflict | known), "Unreported Score"] = 50
    out.loc[(ls == "HIGH_CONFIDENCE_NOT_FOUND") & ~reported_mask & ~(conflict | known) & (fam == 0), "Unreported Score"] = 100
    out.loc[(ls == "MEDIUM_CONFIDENCE_NOT_FOUND") & ~reported_mask, "Unreported Score"] = 80
    out["Unreported Status"] = "LIKELY_UNREPORTED"
    out.loc[reported_mask, "Unreported Status"] = "REPORTED"
    out.loc[conflict | known, "Unreported Status"] = "PRIOR_CONFLICT"
    out.loc[(fam > 0) & ~(reported_mask | conflict | known), "Unreported Status"] = "FAMILY_REVIEW_NEEDED"
    out.loc[out["Unreported Score"] == 100, "Unreported Status"] = "CLEAN_UNREPORTED"
    st = pd.to_numeric(_series(out, "Stability", ""), errors="coerce")
    out["Stability Grade"] = "UNKNOWN"
    out.loc[st <= 0.05, "Stability Grade"] = "EXCELLENT"; out.loc[(st > 0.05) & (st <= 0.10), "Stability Grade"] = "GOOD"; out.loc[(st > 0.10) & (st <= 0.30), "Stability Grade"] = "MODERATE"; out.loc[st > 0.30, "Stability Grade"] = "POOR"
    bg = pd.to_numeric(_series(out, "Band Gap (eV)", ""), errors="coerce")
    out["Band Gap Grade"] = "UNKNOWN"
    out.loc[bg == 0, "Band Gap Grade"] = "METALLIC_OR_ZERO"; out.loc[(bg >= 0.5) & (bg <= 1.5), "Band Gap Grade"] = "IDEAL"; out.loc[((bg >= 0.1) & (bg < 0.5)) | ((bg > 1.5) & (bg <= 2.0)), "Band Gap Grade"] = "GOOD"; out.loc[(bg > 2.0) & (bg <= 3.0), "Band Gap Grade"] = "MODERATE"; out.loc[bg > 3.0, "Band Gap Grade"] = "POOR"
    out["Reported Paper Link"] = out.apply(lambda r: make_paper_link(r.get("paper_link_if_reported", "") or r.get("prior_best_url", "") or r.get("best_url", ""), r.get("prior_best_doi", "") or r.get("best_doi", "")), axis=1)
    out["Toxicity / Practicality"] = _series(out, "practicality_tier", "").map({"PRACTICAL_PRIORITY": "PRACTICAL", "EXPENSIVE_RARE_REVIEW": "EXPENSIVE_RARE", "TOXICITY_REVIEW": "TOXIC", "RADIOACTIVE_REVIEW": "RADIOACTIVE", "HIGHLY_IMPRACTICAL": "HIGHLY_IMPRACTICAL"}).fillna("PRACTICAL")
    for src, label in [("highly_toxic_elements", "TOXIC"), ("expensive_rare_elements", "EXPENSIVE_RARE"), ("radioactive_elements", "RADIOACTIVE")]:
        vals = _series(out, src, "").fillna("").astype(str).str.strip()
        out.loc[vals != "", "Toxicity / Practicality"] = out.loc[vals != "", "Toxicity / Practicality"] + ": " + vals[vals != ""]
    out["Final Decision"] = "MANUAL_REVIEW_REQUIRED"
    out["Final Reason"] = "Not enough confidence for direct recommendation."
    out.loc[out["Unreported Status"] == "REPORTED", ["Final Decision", "Final Reason"]] = ["REPORTED_REJECT", "Reported or formula-equivalent prior evidence found; do not claim novelty."]
    out.loc[out["Unreported Status"] == "PRIOR_CONFLICT", ["Final Decision", "Final Reason"]] = ["MANUAL_REVIEW_REQUIRED", "Prior conflict or known reported composition found; manual citation check required."]
    out.loc[out["Unreported Status"] == "FAMILY_REVIEW_NEEDED", ["Final Decision", "Final Reason"]] = ["MANUAL_REVIEW_REQUIRED", "Family-level literature exists; verify exact material before novelty claim."]
    out.loc[out["Toxicity / Practicality"].str.contains("HIGHLY_IMPRACTICAL|RADIOACTIVE", na=False), ["Final Decision", "Final Reason"]] = ["TOXICITY_REJECT", "Radioactive or highly impractical element present."]
    out.loc[out["Toxicity / Practicality"].str.contains("TOXIC", na=False), ["Final Decision", "Final Reason"]] = ["TOXICITY_REJECT", "Toxic/problematic element present."]
    out.loc[st > 0.30, ["Final Decision", "Final Reason"]] = ["LOW_STABILITY_REJECT", "Likely unreported but stability is too weak for first-priority candidate."]
    out.loc[out["Toxicity / Practicality"].str.contains("EXPENSIVE_RARE", na=False), ["Final Decision", "Final Reason"]] = ["BACKUP_ONLY", "Potentially useful but contains expensive/rare element; use as backup."]
    best_mask = (out["Unreported Status"] == "CLEAN_UNREPORTED") & (st <= 0.10) & (_series(out, "practicality_tier", "") == "PRACTICAL_PRIORITY") & (bg >= 0.1) & (bg <= 2.0) & (out["Reported Paper Link"] == "")
    out.loc[best_mask, ["Final Decision", "Final Reason"]] = ["BEST_NOVEL_CANDIDATE", "Clean unreported candidate with good stability, practical elements, and suitable band gap."]
    good_mask = out["Final Decision"].eq("MANUAL_REVIEW_REQUIRED") & out["Unreported Status"].isin(["CLEAN_UNREPORTED", "LIKELY_UNREPORTED"]) & (st <= 0.20) & (_series(out, "practicality_tier", "") == "PRACTICAL_PRIORITY") & (bg >= 0.1) & (bg <= 3.0)
    out.loc[good_mask, ["Final Decision", "Final Reason"]] = ["GOOD_NOVEL_CANDIDATE", "Promising candidate; manual literature verification recommended."]
    out["_dp"] = pd.Categorical(out["Final Decision"], categories=FINAL_DECISION_SIMPLE_PRIORITY, ordered=True)
    out["_tox"] = out["Toxicity / Practicality"].str.extract(r"^(PRACTICAL|EXPENSIVE_RARE|TOXIC|RADIOACTIVE|HIGHLY_IMPRACTICAL)")[0].map({"PRACTICAL":0,"EXPENSIVE_RARE":1,"TOXIC":2,"RADIOACTIVE":3,"HIGHLY_IMPRACTICAL":4}).fillna(9)
    out["_bg"] = out["Band Gap Grade"].map({"IDEAL":0,"GOOD":1,"MODERATE":2,"POOR":3,"METALLIC_OR_ZERO":4,"UNKNOWN":5}).fillna(5)
    out["_bgclose"] = (bg - 1.0).abs()
    out["_linkblank"] = out["Reported Paper Link"].fillna("").eq("")
    out = out.sort_values(["_dp","Unreported Score","Stability","_tox","_bg","_bgclose","_linkblank"], ascending=[True,False,True,True,True,True,False]).reset_index(drop=True)
    out["Final Rank"] = range(1, len(out) + 1)
    return out


def _build_validation_checks(df: pd.DataFrame) -> pd.DataFrame:
    best = df[df["Final Decision"] == "BEST_NOVEL_CANDIDATE"].copy()
    return pd.DataFrame([
        {"check": "number of BEST_NOVEL_CANDIDATE rows", "value": int((df["Final Decision"] == "BEST_NOVEL_CANDIDATE").sum())},
        {"check": "number of GOOD_NOVEL_CANDIDATE rows", "value": int((df["Final Decision"] == "GOOD_NOVEL_CANDIDATE").sum())},
        {"check": "number of MANUAL_REVIEW_REQUIRED rows", "value": int((df["Final Decision"] == "MANUAL_REVIEW_REQUIRED").sum())},
        {"check": "number of BACKUP_ONLY rows", "value": int((df["Final Decision"] == "BACKUP_ONLY").sum())},
        {"check": "number of REPORTED_REJECT rows", "value": int((df["Final Decision"] == "REPORTED_REJECT").sum())},
        {"check": "warning: BEST has link", "value": int((best["Reported Paper Link"].fillna("") != "").sum())},
        {"check": "warning: BEST stability > 0.10", "value": int((pd.to_numeric(best["Stability"], errors="coerce") > 0.10).sum())},
        {"check": "warning: BEST toxicity not PRACTICAL", "value": int((~best["Toxicity / Practicality"].fillna("").str.startswith("PRACTICAL")).sum())},
        {"check": "warning: BEST band gap outside 0.1-2.0", "value": int(((pd.to_numeric(best["Band Gap (eV)"], errors="coerce") < 0.1) | (pd.to_numeric(best["Band Gap (eV)"], errors="coerce") > 2.0)).sum())},
        {"check": "warning: BEST unreported score < 100", "value": int((best["Unreported Score"] < 100).sum())},
    ])


def export_outputs(df_all: pd.DataFrame, output_dir: Path):
    output_dir.mkdir(parents=True, exist_ok=True)
    df_all.to_csv(output_dir / "05_all_hits_audit.csv", index=False)
    ranked_not_found = df_all[df_all["Automated Status"] == "not_found_after_protocol"].copy()
    if "Unreported Confidence Score" not in ranked_not_found.columns:
        ranked_not_found["Unreported Confidence Score"] = 0
    ranked_not_found.sort_values("Unreported Confidence Score", ascending=False).to_csv(output_dir / "02_ranked_not_found_after_protocol.csv", index=False)
    novelty = df_all.get("novelty_confidence_tier", pd.Series([""] * len(df_all)))
    practical = df_all.get("practicality_tier", pd.Series(["PRACTICAL_PRIORITY"] * len(df_all)))

    df_all[novelty == "HIGH_CONFIDENCE_UNREPORTED"].to_csv(output_dir / "02_high_confidence_unreported.csv", index=False)
    df_all[novelty == "MEDIUM_CONFIDENCE_UNREPORTED"].to_csv(output_dir / "03_medium_confidence_unreported.csv", index=False)
    df_all[novelty == "AMBIGUOUS_REVIEW_REQUIRED"].to_csv(output_dir / "04_ambiguous_manual_review.csv", index=False)
    df_all[df_all["Automated Status"] == "reported_dft"].to_csv(output_dir / "05_reported_dft.csv", index=False)
    df_all[df_all["Automated Status"] == "reported_non_dft"].to_csv(output_dir / "06_reported_non_dft.csv", index=False)
    df_all.to_csv(output_dir / "07_all_hits_audit.csv", index=False)
    df_all.to_csv(output_dir / "08_search_coverage_report.csv", index=False)
    df_all[practical == "PRACTICAL_PRIORITY"].to_csv(output_dir / "09_practical_priority_candidates.csv", index=False)
    df_all[practical != "PRACTICAL_PRIORITY"].to_csv(output_dir / "10_radioactive_or_impractical_candidates.csv", index=False)

    order = ["TOP_RESEARCH_PRIORITY", "HIGH_RESEARCH_PRIORITY", "MEDIUM_RESEARCH_PRIORITY", "LOW_RESEARCH_PRIORITY", "DEFER"]
    ranking = df_all.copy()
    tier_vals = ranking["final_material_priority_tier"] if "final_material_priority_tier" in ranking.columns else pd.Series(["DEFER"] * len(ranking))
    ranking["_tier"] = pd.Categorical(tier_vals, categories=order, ordered=True)
    for col in ["final_selection_score", "keypaper_depth_score", "practicality_score", "stability_score", "literature_risk_penalty"]:
        if col not in ranking.columns:
            ranking[col] = 0
    ranking = ranking.sort_values(["_tier", "final_selection_score", "keypaper_depth_score", "practicality_score", "stability_score", "literature_risk_penalty"], ascending=[True, False, True, False, False, True]).drop(columns=["_tier"])
    ranking.head(10).to_csv(output_dir / "11_top10_final_research_candidates.csv", index=False)

    xlsx_path = output_dir / "01_priority_unreported_candidates.xlsx"
    try:
        with pd.ExcelWriter(xlsx_path) as xw:
            ranking.to_excel(xw, index=False)
    except ModuleNotFoundError:
        xlsx_path.write_bytes(b"")
